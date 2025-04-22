import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def importer_donnees_automatique():
    # Chemins des fichiers (à adapter si nécessaire)
    fichier_source = "depenses_detailles.xlsx"
    fichier_cible = "Mathis - Dépenses 2025.xlsx"
    fichier_backup = "Mathis - Dépenses 2025_backup.xlsx"
    
    # Dictionnaire des catégories et leurs lignes de départ
    categories = {
        "Frais scolaire" : 22, 
        "Automobile & transports en commun" : 41,
        "Linge, manteaux, souliers et bottes" : 71,
        "Sport" : 85,
        "Hygiène" : 97,
        "Soins de santé (physio, dentiste, optométriste, etc.)" : 108,
        "Épicerie" : 119,
        "Restaurant" : 146,
        "Logement, meubles, articles ménagers" : 208,
        "Divers" : 188,
        "Électronique" : 236,
        "Cellulaire" : 6,
        "Billets" : 263
    }
    
    try:
        # Création d'un backup du fichier cible
        if os.path.exists(fichier_cible):
            os.replace(fichier_cible, fichier_backup)
        
        # Chargement des fichiers
        df_source = pd.read_excel(fichier_source)
        classeur = openpyxl.load_workbook(fichier_backup)
        onglet = classeur["Dépenses 2025"]
        
        # Journal des modifications
        modifications = []
        
        def trouver_ligne_vide(onglet, ligne_depart):
            """Trouve la prochaine ligne vide dans une catégorie"""
            ligne = ligne_depart
            while onglet.cell(row=ligne, column=1).value not in [None, "ARRÊT"]:
                ligne += 1
            return ligne
        
        for _, transaction in df_source.iterrows():
            categorie = transaction["Catégorie"]
            if categorie in categories:
                ligne = trouver_ligne_vide(onglet, categories[categorie])
                
                # Vérification des doublons
                existe_deja = False
                for row in range(categories[categorie], ligne):
                    if (str(onglet.cell(row=row, column=3).value).strip() == str(transaction["Description"]).strip() and 
                        float(onglet.cell(row=row, column=6).value or 0) == float(transaction["Après taxes"])):
                        existe_deja = True
                        break
                
                if not existe_deja:
                    # Écriture des données
                    onglet.cell(row=ligne, column=1, value=transaction["Date (A/M/J)"])
                    onglet.cell(row=ligne, column=2, value=int(transaction["Qté"]))
                    onglet.cell(row=ligne, column=3, value=transaction["Description"])
                    onglet.cell(row=ligne, column=4, value=0)  # Prix unitaire
                    onglet.cell(row=ligne, column=6, value=float(transaction["Après taxes"]))
                    
                    modifications.append(
                        f"Ajouté à {categorie}, ligne {ligne}: "
                        f"{transaction['Description']} - {transaction['Après taxes']}$"
                    )
        
        # Sauvegarde des modifications
        classeur.save(fichier_cible)
        
        # Création d'un log
        with open("import_log.txt", "w") as f:
            if modifications:
                f.write("=== TRANSACTIONS AJOUTÉES ===\n")
                f.write("\n".join(modifications))
                f.write(f"\n\nTotal: {len(modifications)} transactions ajoutées")
            else:
                f.write("Aucune nouvelle transaction à ajouter")
        
        print("Importation terminée. Voir import_log.txt pour les détails.")
        
    except Exception as e:
        print(f"ERREUR: {str(e)}")
        if os.path.exists(fichier_backup):
            os.replace(fichier_backup, fichier_cible)
        print("Une sauvegarde a été restaurée")
